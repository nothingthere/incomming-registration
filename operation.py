#!/usr/bin/python3
# excel.py
# Author: Claudio <3261958605@qq.com>
# Created: 2018-04-06 16:04:57
# Code:
'''
对Excel表单的操作
'''

#
# 获取文件夹中的所有文件信息
#
import datetime
import openpyxl as xls
import openpyxl.styles as Styles
import os
import os.path
import time
from openpyxl.styles.borders import Border, Side

# 获取文件夹中的文件信息
# 结果：[[文件名,对应时间],...]并按时间排序。


def timestamp2string(stamp):
    dt = datetime.date.fromtimestamp(stamp)
    return '{}.{:02}.{:02}'.format(dt.year, dt.month, dt.day)


def get_files(direc):
    result = []
    for entry in os.scandir(direc):
        filepath = os.path.splitext(entry.name)[0]
        filetime = os.stat(entry).st_ctime
        # print(filepath, filetime)
        result.append([filepath, filetime])

    import operator
    k = operator.itemgetter(1)
    result = sorted(result, key=k)

    result = map(lambda e: [e[0], timestamp2string(e[1])], result)
    return result

#
# 读写Excel文件
#


"""
1 长山收费站2018年2月收文登记薄
2
3 序号 	收文时间	文 件 名 称 / 标 题	发件单位/部门	文件分类	资料类型	备注
4:开始有内容
"""

FAMILY1 = "DejaVu Sans"  # "方正小标宋简体"
FAMILY2 = "DejaVu Sans"  # "仿宋_GB2312"
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


def write_headers(sheet):
    print(sheet.title)
    rd = sheet.row_dimensions
    print(rd[0].height)
    sheet.merge_cells("A1:G1", start_row=1, end_row=1,
                      start_column=1, end_column=7)
    title = sheet.cell(row=1, column=1)
    title.font = Styles.Font(name=FAMILY1, size=25, bold=False)
    title.alignment = Styles.Alignment(horizontal="center", vertical="center")

    title.value = "长山收费站{}年{}收文登记薄".format(
        datetime.datetime.fromtimestamp(time.time()).year, sheet.title)

    sheet.merge_cells("A2:G2", start_row=2, end_row=2,
                      start_column=1, end_column=7)

    for i, item in enumerate(["序号", "收文时间", "文件名称/标题", "发件单位/部门",
                              "文件分类", "资料类型", "备注"]):
        cell = sheet.cell(row=3, column=i+1)
        cell.alignment = Styles.Alignment(
            horizontal="center", vertical="center")
        cell.border = thin_border
        cell.value = item


def write_items(sheet, files, row_start):
    font12 = Styles.Font(name=FAMILY2, size=12,)
    font14 = Styles.Font(name=FAMILY2, size=14)
    align1 = Styles.Alignment(horizontal="center", vertical="center")
    align2 = Styles.Alignment(horizontal="left", vertical="center")

    for f in files:
        for col in range(1, 7):
            sheet.cell(row=row_start, column=col).border = thin_border

        cell = sheet.cell(row=row_start, column=1)
        cell.font = font12
        cell.alignment = align2
        cell.value = row_start-3

        cell = sheet.cell(row=row_start, column=2)
        cell.font = font14
        cell.alignment = align2

        cell.value = f[1]

        cell = sheet.cell(row=row_start, column=3)
        cell.font = font14
        cell.alignment = align2
        cell.value = f[0]

        row_start += 1


def dump_files(direc, excel):
    """
    direc：文件夹名称兼表单名称
    excel:
    将files写入excel中sheet表单中。
    """
    book = xls.load_workbook(excel)
    all_files = list(get_files(direc))
    registed_files = []
    row_start_append = 4
    sheetname = os.path.basename(direc)
    # 是否存在对应表单
    # 若不存在则创建
    # 若存在则获取已注册文件名
    # 并获取开始插入的行
    if not sheetname in book.sheetnames:
        sheet = book.create_sheet(sheetname)
        write_headers(sheet)

    else:
        sheet = book[sheetname]
        for r, _row in enumerate(sheet.rows):
            # print(r)
            if r <= 2:
                continue
            cell = sheet.cell(row=r+1, column=3)
            cell_value = cell.value
            # print(r+1, cell_value)
            if cell_value:
                registed_files.append(cell_value.strip())
            else:
                break
        row_start_append = r+2

    # 获取需插入的文件信息

    unregisted_files = list(
        filter(lambda f: f[0] not in registed_files, all_files))

    # print("\n所有文件:")
    # for f in all_files:
    #     print(f)
    # print("\n已经注册文件：")
    # for f in registed_files:
    #     print(f)
    # print("\n待写入文件：")
    # for f in unregisted_files:
    #     print(f)


#
# 测试
#

    # # 开始写入
    write_items(sheet, unregisted_files, row_start_append)
    # for f in unregisted_files:
    #     # print(row_start_append-3, f[1], f[0])
    #     sheet.cell(row=row_start_append, column=1, value=row_start_append-3)
    #     sheet.cell(row=row_start_append, column=2, value=f[1])
    #     sheet.cell(row=row_start_append, column=3, value=f[0])
    #     row_start_append += 1

    book.save(excel)


if __name__ == '__main__':
    dump_files('3月', "result.xlsx")
    print('Done!')
